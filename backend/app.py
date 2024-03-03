from flask import Flask, request, jsonify
import json;
import database_connection
import translate
app = Flask(__name__)
from flask_cors import CORS
from flask import abort
from googletrans import Translator
import mysql.connector
import os
from docx import Document
from io import BytesIO
from werkzeug.utils import secure_filename
import openpyxl
from flask import send_file, make_response
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, timedelta
from mysql.connector import pooling
from dateutil.relativedelta import relativedelta
import pandas as pd

CORS(app)

# create a connection pool
mydb = mysql.connector.pooling.MySQLConnectionPool(
    pool_name="my_pool",
    pool_size=32,
    user='root',
    password='',
    host='localhost',
    database='translationapp'
)



google_translate_languages = ['Language Detected Automatically',    'Afrikaans', 'Albanian', 'Amharic', 'Arabic', 'Armenian', 'Azerbaijani',    'Basque', 'Belarusian', 'Bengali', 'Bosnian', 'Bulgarian', 'Catalan',    'Cebuano', 'Chichewa', 'Chinese (Simplified)', 'Chinese (Traditional)',    'Corsican', 'Croatian', 'Czech', 'Danish', 'Dutch', 'English', 'Esperanto',    'Estonian', 'Filipino', 'Finnish', 'French', 'Frisian', 'Galician', 'Georgian',    'German', 'Greek', 'Gujarati', 'Haitian Creole', 'Hausa', 'Hawaiian', 'Hebrew',    'Hindi', 'Hmong', 'Hungarian', 'Icelandic', 'Igbo', 'Indonesian', 'Irish',    'Italian', 'Japanese', 'Javanese', 'Kannada', 'Kazakh', 'Khmer', 'Kinyarwanda',    'Korean', 'Kurdish (Kurmanji)', 'Kyrgyz', 'Lao', 'Latin', 'Latvian', 'Lithuanian',    'Luxembourgish', 'Macedonian', 'Malagasy', 'Malay', 'Malayalam', 'Maltese',    'Maori', 'Marathi', 'Mongolian', 'Myanmar (Burmese)', 'Nepali', 'Norwegian',    'Odia (Oriya)', 'Pashto', 'Persian', 'Polish', 'Portuguese', 'Punjabi', 'Romanian',    'Russian', 'Samoan', 'Scots Gaelic', 'Serbian', 'Sesotho', 'Shona', 'Sindhi',    'Sinhala', 'Slovak', 'Slovenian', 'Somali', 'Spanish', 'Sundanese', 'Swahili',    'Swedish', 'Tajik', 'Tamil', 'Tatar', 'Telugu', 'Thai', 'Turkish', 'Turkmen',    'Ukrainian', 'Urdu', 'Uyghur', 'Uzbek', 'Vietnamese', 'Welsh', 'Xhosa', 'Yiddish',    'Yoruba', 'Zulu']
language_codes = ['auto',
    'af', 'sq', 'am', 'ar', 'hy', 'az', 'eu', 'be', 'bn', 'bs',
    'bg', 'ca', 'ceb', 'ny', 'zh-CN', 'zh-TW', 'co', 'hr', 'cs',
    'da', 'nl', 'en', 'eo', 'et', 'tl', 'fi', 'fr', 'fy', 'gl',
    'ka', 'de', 'el', 'gu', 'ht', 'ha', 'haw', 'iw', 'hi', 'hmn',
    'hu', 'is', 'ig', 'id', 'ga', 'it', 'ja', 'jw', 'kn', 'kk',
    'km', 'rw', 'ko', 'ku', 'ky', 'lo', 'la', 'lv', 'lt', 'lb',
    'mk', 'mg', 'ms', 'ml', 'mt', 'mi', 'mr', 'mn', 'my', 'ne',
    'no', 'or', 'ps', 'fa', 'pl', 'pt', 'pa', 'ro', 'ru', 'sm',
    'gd', 'sr', 'st', 'sn', 'sd', 'si', 'sk', 'sl', 'so', 'es',
    'su', 'sw', 'sv', 'tg', 'ta', 'tt', 'te', 'th', 'tr', 'tk',
    'uk', 'ur', 'ug', 'uz', 'vi', 'cy', 'xh', 'yi', 'yo', 'zu']

#----------------------------------------------------------------------------#
#----------------------------- L I S T E N E R S ----------------------------#
#----------------------------------------------------------------------------#

# Ruta pentru LOGIN
@app.route('/login', methods=['POST'])
def login():
    
    data = request.json
    email = data['email']
    password = data['password']
    print(email, password)
    login_result = database_connection.login(email, password)
    
    if login_result == 1:
         return jsonify({'username': email, 'password': password})
    else:
        abort(409, "Email or password incorrect")
    



@app.route('/register', methods=['POST'])
def register():
    
    data = request.json
    email = data['email']
    password = data['password']
    country = data['country']
    print(email, password)
    register_result = database_connection.register(email, password, country)
    
    if register_result == 1:
         return jsonify({'username': email, 'password': password})
    else:
        abort(409, "User already exists in the database.")


@app.route('/translate', methods=['POST'])
def translateText():
    data = request.json
    text = data.get('OriginalText')
    textFrom = data.get('TargetLanguage')
    textTo = data.get('SourceLanguage')
    if(text==""):
        return "";
    print(text, textFrom, textTo)
    if text:
        translator = Translator()
        translated = translator.translate(text, textTo, textFrom)
    
        return translated.text

    else:
        return jsonify({'error': 'Text parameter missing or empty'})


from datetime import date

from datetime import date
from flask import jsonify

@app.route('/insertTranslate', methods=['POST'])
def insertTranslate():
    data = request.json
    text = data.get('OriginalText')
    textFrom = data.get('TargetLanguage')
    textTo = data.get('SourceLanguage')
    user_id = data.get("UserID")
    translatedText = data.get("translatedText")
    
    if text:        
        conn = mydb.get_connection()
        cursor = conn.cursor()
        
        # Get the current date
        current_date = date.today().isoformat()
        
        insert_query = "INSERT INTO translation (UserID, OriginalText, TranslatedText, SourceLanguage, TargetLanguage, Date) VALUES (%s, %s, %s, %s, %s, %s)"
        insert_values = (user_id, text, translatedText, textFrom, textTo, current_date)
        
        cursor.execute(insert_query, insert_values)
        conn.commit()
        
        cursor.close()
        conn.close()
        
        return jsonify({'message': 'Translation inserted successfully.'})
    
    return jsonify({'error': 'Invalid data provided.'}), 400


@app.route('/returnUID', methods=['POST'])
def returnUID():
    data = request.json
    email = data.get("email")
    if email:
        conn = mydb.get_connection()
        cursor = conn.cursor()
        return_query = f"""
            SELECT user.userID, iso_country_codes.code
            FROM user
            JOIN iso_country_codes ON user.country = iso_country_codes.name
            WHERE user.UserEmail = '{email}'
        """

    cursor.execute(return_query);
    row = cursor.fetchone()
    cursor.close()
    conn.close()
    if row:
        user_id, country_code = row
        return jsonify({"userID": user_id, "country_code": country_code})
    else:
        return jsonify({})

        
@app.route('/returnHistory', methods=['POST'])
def returnHistory():
    google_translate_languages = ['Language Detected Automatically',    'Afrikaans', 'Albanian', 'Amharic', 'Arabic', 'Armenian', 'Azerbaijani',    'Basque', 'Belarusian', 'Bengali', 'Bosnian', 'Bulgarian', 'Catalan',    'Cebuano', 'Chichewa', 'Chinese (Simplified)', 'Chinese (Traditional)',    'Corsican', 'Croatian', 'Czech', 'Danish', 'Dutch', 'English', 'Esperanto',    'Estonian', 'Filipino', 'Finnish', 'French', 'Frisian', 'Galician', 'Georgian',    'German', 'Greek', 'Gujarati', 'Haitian Creole', 'Hausa', 'Hawaiian', 'Hebrew',    'Hindi', 'Hmong', 'Hungarian', 'Icelandic', 'Igbo', 'Indonesian', 'Irish',    'Italian', 'Japanese', 'Javanese', 'Kannada', 'Kazakh', 'Khmer', 'Kinyarwanda',    'Korean', 'Kurdish (Kurmanji)', 'Kyrgyz', 'Lao', 'Latin', 'Latvian', 'Lithuanian',    'Luxembourgish', 'Macedonian', 'Malagasy', 'Malay', 'Malayalam', 'Maltese',    'Maori', 'Marathi', 'Mongolian', 'Myanmar (Burmese)', 'Nepali', 'Norwegian',    'Odia (Oriya)', 'Pashto', 'Persian', 'Polish', 'Portuguese', 'Punjabi', 'Romanian',    'Russian', 'Samoan', 'Scots Gaelic', 'Serbian', 'Sesotho', 'Shona', 'Sindhi',    'Sinhala', 'Slovak', 'Slovenian', 'Somali', 'Spanish', 'Sundanese', 'Swahili',    'Swedish', 'Tajik', 'Tamil', 'Tatar', 'Telugu', 'Thai', 'Turkish', 'Turkmen',    'Ukrainian', 'Urdu', 'Uyghur', 'Uzbek', 'Vietnamese', 'Welsh', 'Xhosa', 'Yiddish',    'Yoruba', 'Zulu']
    language_codes = ['auto',
    'af', 'sq', 'am', 'ar', 'hy', 'az', 'eu', 'be', 'bn', 'bs',
    'bg', 'ca', 'ceb', 'ny', 'zh-CN', 'zh-TW', 'co', 'hr', 'cs',
    'da', 'nl', 'en', 'eo', 'et', 'tl', 'fi', 'fr', 'fy', 'gl',
    'ka', 'de', 'el', 'gu', 'ht', 'ha', 'haw', 'iw', 'hi', 'hmn',
    'hu', 'is', 'ig', 'id', 'ga', 'it', 'ja', 'jw', 'kn', 'kk',
    'km', 'rw', 'ko', 'ku', 'ky', 'lo', 'la', 'lv', 'lt', 'lb',
    'mk', 'mg', 'ms', 'ml', 'mt', 'mi', 'mr', 'mn', 'my', 'ne',
    'no', 'or', 'ps', 'fa', 'pl', 'pt', 'pa', 'ro', 'ru', 'sm',
    'gd', 'sr', 'st', 'sn', 'sd', 'si', 'sk', 'sl', 'so', 'es',
    'su', 'sw', 'sv', 'tg', 'ta', 'tt', 'te', 'th', 'tr', 'tk',
    'uk', 'ur', 'ug', 'uz', 'vi', 'cy', 'xh', 'yi', 'yo', 'zu']
    data = request.json
    user_id = data.get("UserID")
    if user_id:
        conn = mydb.get_connection()
        cursor = conn.cursor()
        return_query = f"SELECT OriginalText, TranslatedText, SourceLanguage, TargetLanguage FROM translation WHERE UserID = '{user_id}' ORDER BY TranslationID DESC LIMIT 15"
        cursor.execute(return_query)
        rows = cursor.fetchall()  # read all results
        #replace all sourceLanguage and targetLanguage with their names
        rows_with_names = []
        for row in rows:
            row_with_names = []
            row_with_names.append(row[0])
            row_with_names.append(row[1])
            row_with_names.append(google_translate_languages[language_codes.index(row[2])])
            row_with_names.append(google_translate_languages[language_codes.index(row[3])])
            rows_with_names.append(row_with_names)
        cursor.close()
        conn.close()
        return jsonify(rows_with_names)
    else:
        return 0

@app.route('/statisticsTranslateFrom', methods=['POST'])
def translateFrom():
    google_translate_languages = ['Language Detected Automatically',    'Afrikaans', 'Albanian', 'Amharic', 'Arabic', 'Armenian', 'Azerbaijani',    'Basque', 'Belarusian', 'Bengali', 'Bosnian', 'Bulgarian', 'Catalan',    'Cebuano', 'Chichewa', 'Chinese (Simplified)', 'Chinese (Traditional)',    'Corsican', 'Croatian', 'Czech', 'Danish', 'Dutch', 'English', 'Esperanto',    'Estonian', 'Filipino', 'Finnish', 'French', 'Frisian', 'Galician', 'Georgian',    'German', 'Greek', 'Gujarati', 'Haitian Creole', 'Hausa', 'Hawaiian', 'Hebrew',    'Hindi', 'Hmong', 'Hungarian', 'Icelandic', 'Igbo', 'Indonesian', 'Irish',    'Italian', 'Japanese', 'Javanese', 'Kannada', 'Kazakh', 'Khmer', 'Kinyarwanda',    'Korean', 'Kurdish (Kurmanji)', 'Kyrgyz', 'Lao', 'Latin', 'Latvian', 'Lithuanian',    'Luxembourgish', 'Macedonian', 'Malagasy', 'Malay', 'Malayalam', 'Maltese',    'Maori', 'Marathi', 'Mongolian', 'Myanmar (Burmese)', 'Nepali', 'Norwegian',    'Odia (Oriya)', 'Pashto', 'Persian', 'Polish', 'Portuguese', 'Punjabi', 'Romanian',    'Russian', 'Samoan', 'Scots Gaelic', 'Serbian', 'Sesotho', 'Shona', 'Sindhi',    'Sinhala', 'Slovak', 'Slovenian', 'Somali', 'Spanish', 'Sundanese', 'Swahili',    'Swedish', 'Tajik', 'Tamil', 'Tatar', 'Telugu', 'Thai', 'Turkish', 'Turkmen',    'Ukrainian', 'Urdu', 'Uyghur', 'Uzbek', 'Vietnamese', 'Welsh', 'Xhosa', 'Yiddish',    'Yoruba', 'Zulu']
    language_codes = ['auto',
    'af', 'sq', 'am', 'ar', 'hy', 'az', 'eu', 'be', 'bn', 'bs',
    'bg', 'ca', 'ceb', 'ny', 'zh-CN', 'zh-TW', 'co', 'hr', 'cs',
    'da', 'nl', 'en', 'eo', 'et', 'tl', 'fi', 'fr', 'fy', 'gl',
    'ka', 'de', 'el', 'gu', 'ht', 'ha', 'haw', 'iw', 'hi', 'hmn',
    'hu', 'is', 'ig', 'id', 'ga', 'it', 'ja', 'jw', 'kn', 'kk',
    'km', 'rw', 'ko', 'ku', 'ky', 'lo', 'la', 'lv', 'lt', 'lb',
    'mk', 'mg', 'ms', 'ml', 'mt', 'mi', 'mr', 'mn', 'my', 'ne',
    'no', 'or', 'ps', 'fa', 'pl', 'pt', 'pa', 'ro', 'ru', 'sm',
    'gd', 'sr', 'st', 'sn', 'sd', 'si', 'sk', 'sl', 'so', 'es',
    'su', 'sw', 'sv', 'tg', 'ta', 'tt', 'te', 'th', 'tr', 'tk',
    'uk', 'ur', 'ug', 'uz', 'vi', 'cy', 'xh', 'yi', 'yo', 'zu']
    data = request.json
    userID = data.get("UserID")
    if userID:
        conn = mydb.get_connection()
        cursor = conn.cursor()
        return_query = f"SELECT SourceLanguage AS Language, COUNT(DISTINCT TranslationID) AS Appearances FROM translation WHERE userID = '{userID}' GROUP BY SourceLanguage"
        cursor.execute(return_query)
        rows = cursor.fetchall()  # read all results
        rows_with_names = [(google_translate_languages[language_codes.index(code)], count) for code, count in rows]
        cursor.close()
        conn.close()
        return jsonify(rows_with_names)
    else:
        cursor.close()
        return 0


@app.route('/globalStatisticsTranslateFrom', methods=['GET'])
def translateFromGlobal():
    google_translate_languages = ['Language Detected Automatically',    'Afrikaans', 'Albanian', 'Amharic', 'Arabic', 'Armenian', 'Azerbaijani',    'Basque', 'Belarusian', 'Bengali', 'Bosnian', 'Bulgarian', 'Catalan',    'Cebuano', 'Chichewa', 'Chinese (Simplified)', 'Chinese (Traditional)',    'Corsican', 'Croatian', 'Czech', 'Danish', 'Dutch', 'English', 'Esperanto',    'Estonian', 'Filipino', 'Finnish', 'French', 'Frisian', 'Galician', 'Georgian',    'German', 'Greek', 'Gujarati', 'Haitian Creole', 'Hausa', 'Hawaiian', 'Hebrew',    'Hindi', 'Hmong', 'Hungarian', 'Icelandic', 'Igbo', 'Indonesian', 'Irish',    'Italian', 'Japanese', 'Javanese', 'Kannada', 'Kazakh', 'Khmer', 'Kinyarwanda',    'Korean', 'Kurdish (Kurmanji)', 'Kyrgyz', 'Lao', 'Latin', 'Latvian', 'Lithuanian',    'Luxembourgish', 'Macedonian', 'Malagasy', 'Malay', 'Malayalam', 'Maltese',    'Maori', 'Marathi', 'Mongolian', 'Myanmar (Burmese)', 'Nepali', 'Norwegian',    'Odia (Oriya)', 'Pashto', 'Persian', 'Polish', 'Portuguese', 'Punjabi', 'Romanian',    'Russian', 'Samoan', 'Scots Gaelic', 'Serbian', 'Sesotho', 'Shona', 'Sindhi',    'Sinhala', 'Slovak', 'Slovenian', 'Somali', 'Spanish', 'Sundanese', 'Swahili',    'Swedish', 'Tajik', 'Tamil', 'Tatar', 'Telugu', 'Thai', 'Turkish', 'Turkmen',    'Ukrainian', 'Urdu', 'Uyghur', 'Uzbek', 'Vietnamese', 'Welsh', 'Xhosa', 'Yiddish',    'Yoruba', 'Zulu']
    language_codes = ['auto',
    'af', 'sq', 'am', 'ar', 'hy', 'az', 'eu', 'be', 'bn', 'bs',
    'bg', 'ca', 'ceb', 'ny', 'zh-CN', 'zh-TW', 'co', 'hr', 'cs',
    'da', 'nl', 'en', 'eo', 'et', 'tl', 'fi', 'fr', 'fy', 'gl',
    'ka', 'de', 'el', 'gu', 'ht', 'ha', 'haw', 'iw', 'hi', 'hmn',
    'hu', 'is', 'ig', 'id', 'ga', 'it', 'ja', 'jw', 'kn', 'kk',
    'km', 'rw', 'ko', 'ku', 'ky', 'lo', 'la', 'lv', 'lt', 'lb',
    'mk', 'mg', 'ms', 'ml', 'mt', 'mi', 'mr', 'mn', 'my', 'ne',
    'no', 'or', 'ps', 'fa', 'pl', 'pt', 'pa', 'ro', 'ru', 'sm',
    'gd', 'sr', 'st', 'sn', 'sd', 'si', 'sk', 'sl', 'so', 'es',
    'su', 'sw', 'sv', 'tg', 'ta', 'tt', 'te', 'th', 'tr', 'tk',
    'uk', 'ur', 'ug', 'uz', 'vi', 'cy', 'xh', 'yi', 'yo', 'zu']
    conn = mydb.get_connection()
    cursor = conn.cursor()
    return_query = f"SELECT SourceLanguage AS Language, COUNT(DISTINCT TranslationID) AS Appearances FROM translation GROUP BY SourceLanguage"
    cursor.execute(return_query)
    rows = cursor.fetchall()  # read all results
    rows_with_names = [(google_translate_languages[language_codes.index(code)], count) for code, count in rows]
    cursor.close()
    conn.close()
    return jsonify(rows_with_names)

@app.route('/statisticsTranslateTo', methods=['POST'])
def translateTo():
        google_translate_languages = ['Language Detected Automatically','Afrikaans', 'Albanian', 'Amharic', 'Arabic', 'Armenian', 'Azerbaijani',    'Basque', 'Belarusian', 'Bengali', 'Bosnian', 'Bulgarian', 'Catalan',    'Cebuano', 'Chichewa', 'Chinese (Simplified)', 'Chinese (Traditional)',    'Corsican', 'Croatian', 'Czech', 'Danish', 'Dutch', 'English', 'Esperanto',    'Estonian', 'Filipino', 'Finnish', 'French', 'Frisian', 'Galician', 'Georgian',    'German', 'Greek', 'Gujarati', 'Haitian Creole', 'Hausa', 'Hawaiian', 'Hebrew',    'Hindi', 'Hmong', 'Hungarian', 'Icelandic', 'Igbo', 'Indonesian', 'Irish',    'Italian', 'Japanese', 'Javanese', 'Kannada', 'Kazakh', 'Khmer', 'Kinyarwanda',    'Korean', 'Kurdish (Kurmanji)', 'Kyrgyz', 'Lao', 'Latin', 'Latvian', 'Lithuanian',    'Luxembourgish', 'Macedonian', 'Malagasy', 'Malay', 'Malayalam', 'Maltese',    'Maori', 'Marathi', 'Mongolian', 'Myanmar (Burmese)', 'Nepali', 'Norwegian',    'Odia (Oriya)', 'Pashto', 'Persian', 'Polish', 'Portuguese', 'Punjabi', 'Romanian',    'Russian', 'Samoan', 'Scots Gaelic', 'Serbian', 'Sesotho', 'Shona', 'Sindhi',    'Sinhala', 'Slovak', 'Slovenian', 'Somali', 'Spanish', 'Sundanese', 'Swahili',    'Swedish', 'Tajik', 'Tamil', 'Tatar', 'Telugu', 'Thai', 'Turkish', 'Turkmen',    'Ukrainian', 'Urdu', 'Uyghur', 'Uzbek', 'Vietnamese', 'Welsh', 'Xhosa', 'Yiddish',    'Yoruba', 'Zulu']
        language_codes = ['auto',
        'af', 'sq', 'am', 'ar', 'hy', 'az', 'eu', 'be', 'bn', 'bs',
        'bg', 'ca', 'ceb', 'ny', 'zh-CN', 'zh-TW', 'co', 'hr', 'cs',
        'da', 'nl', 'en', 'eo', 'et', 'tl', 'fi', 'fr', 'fy', 'gl',
        'ka', 'de', 'el', 'gu', 'ht', 'ha', 'haw', 'iw', 'hi', 'hmn',
        'hu', 'is', 'ig', 'id', 'ga', 'it', 'ja', 'jw', 'kn', 'kk',
        'km', 'rw', 'ko', 'ku', 'ky', 'lo', 'la', 'lv', 'lt', 'lb',
        'mk', 'mg', 'ms', 'ml', 'mt', 'mi', 'mr', 'mn', 'my', 'ne',
        'no', 'or', 'ps', 'fa', 'pl', 'pt', 'pa', 'ro', 'ru', 'sm',
        'gd', 'sr', 'st', 'sn', 'sd', 'si', 'sk', 'sl', 'so', 'es',
        'su', 'sw', 'sv', 'tg', 'ta', 'tt', 'te', 'th', 'tr', 'tk',
        'uk', 'ur', 'ug', 'uz', 'vi', 'cy', 'xh', 'yi', 'yo', 'zu']
        data = request.json
        userID = data.get("UserID")
        if userID:
            conn = mydb.get_connection()
            cursor2 = conn.cursor()
            return_query = f"SELECT TargetLanguage AS Language, COUNT(DISTINCT TranslationID) AS Appearances FROM translation WHERE userID = '{userID}' GROUP BY TargetLanguage"
            cursor2.execute(return_query)
            rows = cursor2.fetchall()  # read all results
            rows_with_names = [(google_translate_languages[language_codes.index(code)], count) for code, count in rows]
            print(rows_with_names)
            cursor2.close()
            conn.close()
            print(rows)
            return jsonify(rows_with_names)
        else:
            cursor2.close()
            return 0


@app.route('/globalStatisticsTranslateTo', methods=['GET'])
def translateToGlobal():
        google_translate_languages = ['Language Detected Automatically','Afrikaans', 'Albanian', 'Amharic', 'Arabic', 'Armenian', 'Azerbaijani',    'Basque', 'Belarusian', 'Bengali', 'Bosnian', 'Bulgarian', 'Catalan',    'Cebuano', 'Chichewa', 'Chinese (Simplified)', 'Chinese (Traditional)',    'Corsican', 'Croatian', 'Czech', 'Danish', 'Dutch', 'English', 'Esperanto',    'Estonian', 'Filipino', 'Finnish', 'French', 'Frisian', 'Galician', 'Georgian',    'German', 'Greek', 'Gujarati', 'Haitian Creole', 'Hausa', 'Hawaiian', 'Hebrew',    'Hindi', 'Hmong', 'Hungarian', 'Icelandic', 'Igbo', 'Indonesian', 'Irish',    'Italian', 'Japanese', 'Javanese', 'Kannada', 'Kazakh', 'Khmer', 'Kinyarwanda',    'Korean', 'Kurdish (Kurmanji)', 'Kyrgyz', 'Lao', 'Latin', 'Latvian', 'Lithuanian',    'Luxembourgish', 'Macedonian', 'Malagasy', 'Malay', 'Malayalam', 'Maltese',    'Maori', 'Marathi', 'Mongolian', 'Myanmar (Burmese)', 'Nepali', 'Norwegian',    'Odia (Oriya)', 'Pashto', 'Persian', 'Polish', 'Portuguese', 'Punjabi', 'Romanian',    'Russian', 'Samoan', 'Scots Gaelic', 'Serbian', 'Sesotho', 'Shona', 'Sindhi',    'Sinhala', 'Slovak', 'Slovenian', 'Somali', 'Spanish', 'Sundanese', 'Swahili',    'Swedish', 'Tajik', 'Tamil', 'Tatar', 'Telugu', 'Thai', 'Turkish', 'Turkmen',    'Ukrainian', 'Urdu', 'Uyghur', 'Uzbek', 'Vietnamese', 'Welsh', 'Xhosa', 'Yiddish',    'Yoruba', 'Zulu']
        language_codes = ['auto',
        'af', 'sq', 'am', 'ar', 'hy', 'az', 'eu', 'be', 'bn', 'bs',
        'bg', 'ca', 'ceb', 'ny', 'zh-CN', 'zh-TW', 'co', 'hr', 'cs',
        'da', 'nl', 'en', 'eo', 'et', 'tl', 'fi', 'fr', 'fy', 'gl',
        'ka', 'de', 'el', 'gu', 'ht', 'ha', 'haw', 'iw', 'hi', 'hmn',
        'hu', 'is', 'ig', 'id', 'ga', 'it', 'ja', 'jw', 'kn', 'kk',
        'km', 'rw', 'ko', 'ku', 'ky', 'lo', 'la', 'lv', 'lt', 'lb',
        'mk', 'mg', 'ms', 'ml', 'mt', 'mi', 'mr', 'mn', 'my', 'ne',
        'no', 'or', 'ps', 'fa', 'pl', 'pt', 'pa', 'ro', 'ru', 'sm',
        'gd', 'sr', 'st', 'sn', 'sd', 'si', 'sk', 'sl', 'so', 'es',
        'su', 'sw', 'sv', 'tg', 'ta', 'tt', 'te', 'th', 'tr', 'tk',
        'uk', 'ur', 'ug', 'uz', 'vi', 'cy', 'xh', 'yi', 'yo', 'zu']
        conn = mydb.get_connection()
        cursor2 = conn.cursor()
        return_query = f"SELECT TargetLanguage AS Language, COUNT(DISTINCT TranslationID) AS Appearances FROM translation GROUP BY TargetLanguage"
        cursor2.execute(return_query)
        rows = cursor2.fetchall()  # read all results
        rows_with_names = [(google_translate_languages[language_codes.index(code)], count) for code, count in rows]
        print(rows_with_names)
        cursor2.close()
        conn.close()
        print(rows)
        return jsonify(rows_with_names)

@app.route('/languageOptions', methods=['GET'])
def languageOptions():
    google_translate_languages = ['Afrikaans', 'Albanian', 'Amharic', 'Arabic', 'Armenian', 'Azerbaijani',    'Basque', 'Belarusian', 'Bengali', 'Bosnian', 'Bulgarian', 'Catalan',    'Cebuano', 'Chichewa', 'Chinese (Simplified)', 'Chinese (Traditional)',    'Corsican', 'Croatian', 'Czech', 'Danish', 'Dutch', 'English', 'Esperanto',    'Estonian', 'Filipino', 'Finnish', 'French', 'Frisian', 'Galician', 'Georgian',    'German', 'Greek', 'Gujarati', 'Haitian Creole', 'Hausa', 'Hawaiian', 'Hebrew',    'Hindi', 'Hmong', 'Hungarian', 'Icelandic', 'Igbo', 'Indonesian', 'Irish',    'Italian', 'Japanese', 'Javanese', 'Kannada', 'Kazakh', 'Khmer', 'Kinyarwanda',    'Korean', 'Kurdish (Kurmanji)', 'Kyrgyz', 'Lao', 'Latin', 'Latvian', 'Lithuanian',    'Luxembourgish', 'Macedonian', 'Malagasy', 'Malay', 'Malayalam', 'Maltese',    'Maori', 'Marathi', 'Mongolian', 'Myanmar (Burmese)', 'Nepali', 'Norwegian',    'Odia (Oriya)', 'Pashto', 'Persian', 'Polish', 'Portuguese', 'Punjabi', 'Romanian',    'Russian', 'Samoan', 'Scots Gaelic', 'Serbian', 'Sesotho', 'Shona', 'Sindhi',    'Sinhala', 'Slovak', 'Slovenian', 'Somali', 'Spanish', 'Sundanese', 'Swahili',    'Swedish', 'Tajik', 'Tamil', 'Tatar', 'Telugu', 'Thai', 'Turkish', 'Turkmen',    'Ukrainian', 'Urdu', 'Uyghur', 'Uzbek', 'Vietnamese', 'Welsh', 'Xhosa', 'Yiddish',    'Yoruba', 'Zulu']
    language_codes = [
    'af', 'sq', 'am', 'ar', 'hy', 'az', 'eu', 'be', 'bn', 'bs',
    'bg', 'ca', 'ceb', 'ny', 'zh-CN', 'zh-TW', 'co', 'hr', 'cs',
    'da', 'nl', 'en', 'eo', 'et', 'tl', 'fi', 'fr', 'fy', 'gl',
    'ka', 'de', 'el', 'gu', 'ht', 'ha', 'haw', 'iw', 'hi', 'hmn',
    'hu', 'is', 'ig', 'id', 'ga', 'it', 'ja', 'jw', 'kn', 'kk',
    'km', 'rw', 'ko', 'ku', 'ky', 'lo', 'la', 'lv', 'lt', 'lb',
    'mk', 'mg', 'ms', 'ml', 'mt', 'mi', 'mr', 'mn', 'my', 'ne',
    'no', 'or', 'ps', 'fa', 'pl', 'pt', 'pa', 'ro', 'ru', 'sm',
    'gd', 'sr', 'st', 'sn', 'sd', 'si', 'sk', 'sl', 'so', 'es',
    'su', 'sw', 'sv', 'tg', 'ta', 'tt', 'te', 'th', 'tr', 'tk',
    'uk', 'ur', 'ug', 'uz', 'vi', 'cy', 'xh', 'yi', 'yo', 'zu'
    ]
    data = {'google_translate_languages': google_translate_languages, 'language_codes': language_codes}
    json_data = json.dumps(data)
    return json_data


@app.route('/create_folder', methods=['POST'])
def create_folder():
    userID = request.json['userID']
    directory = str(userID)
    parent_dir = "C:/Users/alexk/proiect-colectiv/Files/" # Adauga calea unde vrei sa fie create folderele
    path = os.path.join(parent_dir, directory)
    try:
        os.mkdir(path)
        return f"Folder '{userID}' was created successfully."
    except Exception as e:
        return str(e)


from werkzeug.utils import secure_filename

# ...

@app.route('/upload_file', methods=['POST'])
def upload_file():
    userID = request.form.get('userID')
    file = request.files.get('file')

    if not userID:
        return 'Missing userID field', 400

    if not file:
        return 'Missing file', 400

    filename = secure_filename(file.filename)
    file_extension = os.path.splitext(filename)[1].lower()
    
    print(file_extension)  # Add this line to print the file extension

    if file_extension == '.docx':
        path = f"C:/Users/alexk/proiect-colectiv/Files/{userID}"
        if not os.path.exists(path):
            os.makedirs(path)
        file.save(os.path.join(path, filename))
    elif file_extension == '.txt':
        text = file.read().decode('utf-8')
        path = f"C:/Users/alexk/proiect-colectiv/Files/{userID}/{filename}"
        with open(path, 'w') as text_file:
            text_file.write(text)
    elif file_extension == '.xlsx':
        path = f"C:/Users/alexk/proiect-colectiv/Files/{userID}"
        if not os.path.exists(path):
            os.makedirs(path)
        file.save(os.path.join(path, filename))
    else:
        return 'Invalid file format', 400


    return 'File uploaded and processed successfully'



@app.route('/get_file_names', methods=['POST'])
def get_folder_files():
    data = request.json
    userID = data.get("UserID")  # Use lowercase 'userID' instead of 'UserID'
    print(userID)
    parent_dir = "C:/Users/alexk/proiect-colectiv/Files/"
    path = os.path.join(parent_dir, str(userID))

    if not os.path.exists(path):
        return f"Folder '{userID}' does not exist."

    file_names = os.listdir(path)
    file_names.sort(key=lambda x: os.path.getmtime(os.path.join(path, x)), reverse=True)
    return jsonify(file_names)


@app.route('/get_file_content', methods=['POST'])
def read_text_file():
    translator = Translator()
    data = request.json
    file_name = data.get("file_name")
    userID = data.get("userID")
    sourceLanguage = data.get("sourceLanguage")
    targetLanguage = data.get("targetLanguage")
    x = data.get("x")

    try:
        file_path = os.path.abspath(f'C:/Users/alexk/proiect-colectiv/Files/{str(userID)}/{file_name}')
        file_extension = os.path.splitext(file_name)[1]  # Get the file extension
        
        if file_name.endswith(".txt"):
            with open(file_path, 'r') as file:
                text = file.read().lstrip()  # Remove leading blank spaces
        elif file_name.endswith(".docx"):
            with open(file_path, 'rb') as file:
                docx_bytes = file.read()
            
            doc = Document(BytesIO(docx_bytes))
            paragraphs = doc.paragraphs
            text = "\n".join([p.text for p in paragraphs]).lstrip()  # Remove leading blank spaces

        elif file_name.endswith(".xlsx"):
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            rows_data = [", ".join(map(str, row)) for row in rows]
            print(rows_data)  # Add this line to print the rows data

            text = "\n".join(rows_data).lstrip()  # Remove leading blank spaces

        else:
            return jsonify({"error": "Unsupported file type."})

        translated = translator.translate(text, sourceLanguage, targetLanguage)
        
        
        if x==1:
            upload_translated_files(userID, file_extension);

        return jsonify({"original": text, "translated": translated.text})
    
    except FileNotFoundError:
        error_message = f"File '{file_path}' not found."
        print(error_message)
        return jsonify({"error": error_message})

def upload_translated_files(userID, file_extension):
        conn = mydb.get_connection()
        cursor = conn.cursor()
        insert_query = "INSERT INTO files (UserID, FileName) VALUES (%s, %s)"
        insert_values = (userID, file_extension)
        cursor.execute(insert_query, insert_values)
        conn.commit()        
        cursor.close()
        conn.close()

@app.route('/get_file_content_excel', methods=['POST'])
def read_text_file_excel():
    translator = Translator()
    data = request.json
    file_name = data.get("file_name")
    userID = data.get("userID")
    sourceLanguage = data.get("sourceLanguage")
    targetLanguage = data.get("targetLanguage")

    try:
        file_path = os.path.abspath(f'C:/Users/alexk/proiect-colectiv/Files/{str(userID)}/{file_name}')

        if file_name.endswith(".xlsx"):
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

            translated_data = []
            for row in sheet.iter_rows(values_only=True):
                translated_row = [translator.translate(cell, sourceLanguage, targetLanguage).text if cell is not None else None for cell in row]
                translated_data.append(translated_row)

            translated_bytes_io = BytesIO()
            translated_workbook = openpyxl.Workbook()
            translated_sheet = translated_workbook.active

            for translated_row in translated_data:
                translated_sheet.append(translated_row)

            translated_workbook.save(translated_bytes_io)
            translated_bytes_io.seek(0)

            response = make_response(translated_bytes_io.getvalue())
            response.headers['Content-Disposition'] = 'attachment; filename=translated.xlsx'
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            return response

        else:
            return jsonify({"error": "Unsupported file type."})

    except FileNotFoundError:
        error_message = f"File '{file_path}' not found."
        print(error_message)
        return jsonify({"error": error_message})




@app.route('/delete_file', methods=['POST'])
def delete_file():
    userID = request.json.get('userID')
    file_name = request.json.get('file_name')

    if not userID or not file_name:
        return 'Missing userID or file_name field', 400

    file_path = f'C:/Users/alexk/proiect-colectiv/Files/{str(userID)}/{file_name}'

    if os.path.exists(file_path):
        os.remove(file_path)
        return 'File deleted successfully'
    else:
        return 'File not found', 404


from datetime import datetime, timedelta

from datetime import datetime, timedelta

from datetime import datetime, timedelta

@app.route('/translations_per_day', methods=['POST'])
def translations_per_day():
    data = request.json
    user_id = data.get("userID")
    if user_id:
        # Get the current date and calculate the start and end dates for the past 30 days
        today = datetime.now().date()
        start_date = today - timedelta(days=29)
        end_date = today

        # Prepare the query to fetch translation counts for each day
        query = """
        SELECT DATE(translation.Date) AS Date, COUNT(*) AS TranslationCount
        FROM translation
        WHERE translation.UserID = %s AND translation.Date >= %s AND translation.Date <= %s
        GROUP BY DATE(translation.Date)
        ORDER BY DATE(translation.Date)
        """

        conn = mydb.get_connection()
        cursor = conn.cursor()

        # Executing the query and fetching the results
        cursor.execute(query, (user_id, start_date, end_date))
        results = cursor.fetchall()
        cursor.close()
        conn.close()

        # Create a dictionary to store the translation data
        translation_data = {'day': [], 'value': []}

        # Iterate over the past 30 days
        current_date = start_date
        while current_date <= end_date:
            # Check if the current date is present in the results
            found = False
            for row in results:
                if row[0] == current_date:
                    formatted_date = current_date.strftime('%B %d, %Y')
                    translation_data['day'].append(formatted_date)
                    translation_data['value'].append(row[1])
                    found = True
                    break
            if not found:
                # If the current date is not in the results, add it with zero translations
                formatted_date = current_date.strftime('%B %d, %Y')
                translation_data['day'].append(formatted_date)
                translation_data['value'].append(0)
            # Move to the next day
            current_date += timedelta(days=1)

        return jsonify(translation_data)







@app.route('/file_counts', methods=['POST'])
def get_file_counts():
    data = request.json
    user_id = data.get("userID")

    conn = mydb.get_connection()
    cursor = conn.cursor()

    query = """
    SELECT FileName, COUNT(*) AS FileCount
    FROM files
    WHERE UserID = %s
    GROUP BY FileName
    """
    cursor.execute(query, (user_id,))
    results = cursor.fetchall()
    file_counts = {}
    for row in results:
        file_extension = row[0]
        count = row[1]
        file_counts[file_extension] = count

    return jsonify({'data': file_counts})

@app.route('/top_10_countries', methods=['POST'])
def get_top_countries():
    # Establish a connection to the database
    
    conn = mydb.get_connection()
    cursor = conn.cursor()

    # Execute the query to retrieve the top 10 countries with the most translations
    query = """
    SELECT u.country, COUNT(t.TranslationID) AS translation_count
    FROM user u
    JOIN translation t ON u.userID = t.UserID
    GROUP BY u.country
    ORDER BY translation_count DESC
    LIMIT 10
    """
    cursor.execute(query)

    # Fetch all the results
    results = cursor.fetchall()

    # Create a list of dictionaries to store the results
    top_countries = []
    for row in results:
        country = row[0]
        translation_count = row[1]
        top_countries.append({"country": country, "translation_count": translation_count})

    # Close the cursor and database connection
    cursor.close()
    conn.close()

    # Return the result as JSON
    return jsonify({"data": top_countries})

@app.route('/get_file_original_excel', methods=['POST'])
def get_file_original_excel():
    data = request.json
    file_name = data.get("file_name")
    userID = data.get("userID")
    
    file_path = f'C:/Users/alexk/proiect-colectiv/Files/{str(userID)}/{file_name}'

    # Use the file_path to read the Excel file and extract the contents
    # Here's an example of how you can read the file using the pandas library
    df = pd.read_excel(file_path)

    # Convert the DataFrame to an array of objects (list of dictionaries) for easy JSON serialization
    table_data = df.to_dict('records')

    return jsonify(table_data)




if __name__ == '__main__':
    app.run(debug=True)

